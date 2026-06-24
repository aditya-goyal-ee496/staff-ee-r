"""`SupplyDemandSource` backed by the Parity Partners demand-supply workbook (Track B, slice 01).

Each tab carries a title row (`... as of YYYY-MM-DD ...`), a header row, then data rows. This
adapter normalises the Open Roles tab and the three supply tabs (Beach / Rolling Off / New
Joiners) into domain `Role`/`Consultant` value objects. Malformed dates, priorities, or
confidence values fail loudly as `SupplyDemandError` — infrastructure detail never leaks into the
domain, and rows are never silently dropped
(`.claude/principles/hexagonal-architecture.md`, slice 01).
"""

from __future__ import annotations

import re
from collections.abc import Iterator, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from staffeer.adapters.memory_supply_demand import InMemorySupplyDemandSource
from staffeer.domain.errors import SupplyDemandError
from staffeer.domain.models import Consultant, Priority, Role, SupplyState

_AS_OF = re.compile(r"as of (\d{4}-\d{2}-\d{2})")
_CONFIDENCE = {"low": 0.3, "medium": 0.6, "high": 0.9}

_Row = Sequence[Any]


@dataclass(frozen=True)
class _SupplyLayout:
    """Where each `Consultant` field sits in a supply tab (column indices vary per tab)."""

    sheet: str
    state: SupplyState
    name: int
    grade: int
    skills: int
    location: int
    chennai_open: int
    number: int = 0
    available: int | None = None  # date column; None means "available now" (beach → as-of date)
    confidence: int | None = None  # textual confidence column; None means full confidence
    skills_verified: bool = True


_SUPPLY_TABS = (
    _SupplyLayout(
        "Beach", SupplyState.BEACH, name=1, grade=3, skills=4, location=5, chennai_open=6
    ),
    _SupplyLayout(
        "Rolling Off",
        SupplyState.ROLLING_OFF,
        name=1,
        grade=3,
        skills=4,
        location=8,
        chennai_open=9,
        available=6,
        confidence=7,
    ),
    _SupplyLayout(
        "New Joiners",
        SupplyState.NEW_JOINER,
        name=1,
        grade=3,
        skills=4,
        location=6,
        chennai_open=7,
        available=5,
        skills_verified=False,  # CV skills are unverified for new joiners (brief)
    ),
)


class XlsxSupplyDemandSource:
    """Reads roles and consultants from the demand-supply workbook into domain models."""

    def __init__(self, path: str | Path) -> None:
        roles, consultants = _load(Path(path))
        self._delegate = InMemorySupplyDemandSource(roles, consultants)

    def open_roles(self) -> tuple[Role, ...]:
        """All open roles parsed from the workbook."""
        return self._delegate.open_roles()

    def role(self, role_id: str) -> Role:
        """The role with `role_id`, or raise `SupplyDemandError` if unknown."""
        return self._delegate.role(role_id)

    def consultants(self, *states: SupplyState) -> tuple[Consultant, ...]:
        """Consultants filtered by `states` (all when no states are given)."""
        return self._delegate.consultants(*states)


def _load(path: Path) -> tuple[tuple[Role, ...], tuple[Consultant, ...]]:
    """Open the workbook and parse every tab, mapping any load failure to `SupplyDemandError`."""
    try:
        workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except (FileNotFoundError, InvalidFileException, BadZipFile) as exc:
        raise SupplyDemandError(f"cannot read workbook {path}: {exc}") from exc
    try:
        roles = tuple(_parse_roles(_rows(workbook, "Open Roles")))
        consultants = tuple(
            consultant for tab in _SUPPLY_TABS for consultant in _parse_supply(workbook, tab)
        )
    finally:
        workbook.close()
    return roles, consultants


def _rows(workbook: Any, sheet: str) -> list[_Row]:
    """Every row of `sheet` as value tuples, or `SupplyDemandError` if the tab is missing."""
    if sheet not in workbook.sheetnames:
        raise SupplyDemandError(f"workbook is missing the {sheet!r} tab")
    return list(workbook[sheet].iter_rows(values_only=True))


def _data_rows(rows: Sequence[_Row]) -> Iterator[_Row]:
    """Data rows only (skip the title and header rows, and any blank padding rows)."""
    for row in rows[2:]:
        if row and row[0] is not None and str(row[0]).strip():
            yield row


def _parse_roles(rows: Sequence[_Row]) -> Iterator[Role]:
    """Parse the Open Roles tab into `Role`s."""
    for row in _data_rows(rows):
        role_id = str(row[0]).strip()
        location = str(row[6]).strip()
        co_location = _is_yes(row[7])
        yield Role(
            id=role_id,
            title=str(row[1]).strip(),
            required_skills=_split(row[4], ";"),
            start_date=_parse_date(row[5], f"role {role_id} start"),
            location=location,
            priority=_parse_priority(row[8], role_id),
            co_location=co_location,
            chennai_open=co_location and "chennai" in location.lower(),
        )


def _parse_supply(workbook: Any, tab: _SupplyLayout) -> Iterator[Consultant]:
    """Parse one supply tab into `Consultant`s, dating beach availability from the as-of date."""
    rows = _rows(workbook, tab.sheet)
    as_of = _as_of_date(rows, tab.sheet)
    for row in _data_rows(rows):
        consultant_id = f"{tab.state.value}-{int(row[tab.number])}"
        yield Consultant(
            id=consultant_id,
            name=str(row[tab.name]).strip(),
            location=str(row[tab.location]).strip(),
            grade=_optional(row[tab.grade]),
            skills=_split(row[tab.skills], ","),
            state=tab.state,
            chennai_open=_is_yes(row[tab.chennai_open]),
            available_from=as_of
            if tab.available is None
            else _parse_date(row[tab.available], f"{consultant_id} availability"),
            confidence=1.0
            if tab.confidence is None
            else _parse_confidence(row[tab.confidence], consultant_id),
            skills_verified=tab.skills_verified,
        )


def _as_of_date(rows: Sequence[_Row], sheet: str) -> date:
    """The `as of YYYY-MM-DD` date from a tab's title row."""
    title = str(rows[0][0]) if rows and rows[0] else ""
    match = _AS_OF.search(title)
    if not match:
        raise SupplyDemandError(f"{sheet!r} tab is missing an 'as of <date>' title")
    return date.fromisoformat(match.group(1))


def _parse_date(value: Any, context: str) -> date:
    """A `date` from a cell that may hold a date, datetime, or ISO string."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError as exc:
        raise SupplyDemandError(f"{context}: invalid date {value!r}") from exc


def _parse_priority(value: Any, role_id: str) -> Priority:
    """A `Priority` from a `High`/`Medium`/`Low` cell."""
    try:
        return Priority(str(value).strip().lower())
    except ValueError as exc:
        raise SupplyDemandError(f"role {role_id}: invalid priority {value!r}") from exc


def _parse_confidence(value: Any, consultant_id: str) -> float:
    """A confidence weight from a `low`/`medium`/`high` cell."""
    try:
        return _CONFIDENCE[str(value).strip().lower()]
    except KeyError as exc:
        raise SupplyDemandError(f"{consultant_id}: invalid confidence {value!r}") from exc


def _split(value: Any, separator: str) -> tuple[str, ...]:
    """Split a delimited skills cell into trimmed, non-empty skills."""
    if value is None:
        return ()
    return tuple(part.strip() for part in str(value).split(separator) if part.strip())


def _optional(value: Any) -> str | None:
    """A trimmed string, or `None` when the cell is empty."""
    text = str(value).strip() if value is not None else ""
    return text or None


def _is_yes(value: Any) -> bool:
    """True when a Yes/No cell reads yes."""
    return str(value).strip().lower() == "yes"
